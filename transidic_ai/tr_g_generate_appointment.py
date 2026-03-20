import os
import re
import sys
from datetime import datetime, timedelta
from collections import defaultdict
import tkinter as tk
from tkinter import filedialog

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_JUSTIFY

from openpyxl import Workbook, load_workbook

# -----------------------------
# CONFIG
# -----------------------------
WHATSAPP_NUMBER = "+263 785 261 617"

# -----------------------------
# GLOBAL (set after username input)
# -----------------------------
USER_FOLDER = None
ROOT_APPOINTMENTS_FOLDER = None
LOGO_PATH = None

# -----------------------------
# GENERATE APPOINTMENT ID
# -----------------------------
def generate_id(entity_type, counter, client_key):
    date_str = datetime.now().strftime("%Y%m%d")
    prefix = "CS" if entity_type == "cloud" else "CON"
    return f"{client_key}-APT-{date_str}-{prefix}-{counter:03d}"

# -----------------------------
# EXTRACT CLIENT KEY
# -----------------------------
def extract_client_key(file_path):
    with open(file_path, "r", encoding="utf-8") as f:
        content = f.read()

    match = re.search(r"\b[A-Z]{2}-[A-Z0-9]{6}\b", content)
    if match:
        return match.group(0)

    raise Exception("Encrypted client key (e.g. BW-74B35A) not found in invoice.")

# -----------------------------
# PARSE INVOICE TABLE
# -----------------------------
def parse_invoice(file_path):
    cloud_data = defaultdict(lambda: {"tasks": [], "total": 0.0})
    consultant_data = defaultdict(lambda: {"tasks": [], "total": 0.0})

    with open(file_path, "r", encoding="utf-8") as f:
        lines = f.readlines()

    table_started = False
    current_row = None

    for line in lines:
        if "| Item |" in line:
            table_started = True
            continue

        if not table_started or "|" not in line or line.startswith("+"):
            continue

        cols = [c.strip() for c in line.split("|")[1:-1]]
        if len(cols) < 11:
            continue

        item, description, cloud_service, consultant = cols[0], cols[1], cols[2], cols[4]
        cloud_amt, consultant_amt = cols[7], cols[8]

        if item:
            if current_row:
                try:
                    cloud_value = float(current_row["cloud_amt"].replace("$", ""))
                    consultant_value = float(current_row["consultant_amt"].replace("$", ""))
                except:
                    cloud_value, consultant_value = 0.0, 0.0

                cloud_name = current_row["cloud_service"].strip()
                consultant_name = current_row["consultant"].strip()
                task = " ".join(current_row["description"].split())

                cloud_data[cloud_name]["tasks"].append(task)
                cloud_data[cloud_name]["total"] += cloud_value

                consultant_data[consultant_name]["tasks"].append(task)
                consultant_data[consultant_name]["total"] += consultant_value

            current_row = {
                "description": description,
                "cloud_service": cloud_service,
                "consultant": consultant,
                "cloud_amt": cloud_amt,
                "consultant_amt": consultant_amt,
            }

        elif current_row:
            if description:
                current_row["description"] += " " + description
            if cloud_service:
                current_row["cloud_service"] += " " + cloud_service
            if consultant:
                current_row["consultant"] += " " + consultant

    if current_row:
        try:
            cloud_value = float(current_row["cloud_amt"].replace("$", ""))
            consultant_value = float(current_row["consultant_amt"].replace("$", ""))
        except:
            cloud_value, consultant_value = 0.0, 0.0

        cloud_name = current_row["cloud_service"].strip()
        consultant_name = current_row["consultant"].strip()
        task = " ".join(current_row["description"].split())

        cloud_data[cloud_name]["tasks"].append(task)
        cloud_data[cloud_name]["total"] += cloud_value

        consultant_data[consultant_name]["tasks"].append(task)
        consultant_data[consultant_name]["total"] += consultant_value

    return cloud_data, consultant_data

# -----------------------------
# BUILD LETTER CONTENT
# -----------------------------
def build_letter_flowables(entity_name, tasks, total_cost, appointment_id, entity_type, client_key, appointment_time):
    styles = getSampleStyleSheet()

    heading_style = ParagraphStyle(
        "heading",
        parent=styles["Heading3"],
        spaceBefore=8,
        spaceAfter=4
    )

    body_style = ParagraphStyle(
        "body",
        parent=styles["Normal"],
        spaceAfter=10,
        leading=16,
        alignment=TA_JUSTIFY
    )

    entity_label = "Cloud Service" if entity_type == "cloud" else "Consultant"
    elements = []

    deadline = appointment_time + timedelta(hours=48)

    elements.append(Paragraph(f"Client Key: <b>{client_key}</b>", body_style))
    elements.append(Paragraph(f"Appointment Reference: <b>{appointment_id}</b>", body_style))
    elements.append(Paragraph(f"Appointed {entity_label}: <b>{entity_name}</b>", body_style))
    elements.append(Paragraph(f"Appointment Date & Time: <b>{appointment_time.strftime('%Y-%m-%d %H:%M:%S')}</b>", body_style))
    elements.append(Paragraph(f"Deadline (48 hours from appointment): <b>{deadline.strftime('%Y-%m-%d %H:%M:%S')}</b>", body_style))
    elements.append(Spacer(1, 8))

    elements.append(Paragraph(f"Congratulations {entity_name}!", heading_style))

    elements.append(Paragraph(
        "We are delighted to invite you to collaborate with our vibrant and forward-thinking innovation ecosystem. "
        "You have been selected to participate in exciting solution development opportunities within The Constructaxis Smart Infrastructure Ecosystem.",
        body_style
    ))

    elements.append(Spacer(1, 8))
    elements.append(Paragraph("Assigned Tasks", heading_style))

    for idx, task in enumerate(tasks, start=1):
        if task:
            elements.append(Paragraph(f"{idx}. <b>{task}</b>", body_style))

    elements.append(Spacer(1, 8))
    elements.append(Paragraph(f"Total Compensation (Hourly Basis): <b>${total_cost:.2f}</b>", body_style))

    elements.append(Spacer(1, 12))
    elements.append(Paragraph("Kind regards,", body_style))
    elements.append(Paragraph("Constructaxis Cloud Administrator", body_style))
    elements.append(Paragraph("+263 785 261 617", body_style))

    return elements

# -----------------------------
# SAVE PDF
# -----------------------------
def save_pdf(folder, name, appointment_id, client_key, entity_type, flowables):
    safe_name = re.sub(r"[^\w\- ]", "", name).replace(" ", "_")
    safe_type = "Cloud_Service" if entity_type == "cloud" else "Consultant"

    file_path = os.path.join(
        folder,
        f"{appointment_id}_{safe_type}_{safe_name}.pdf"
    )

    doc = SimpleDocTemplate(
        file_path,
        pagesize=A4,
        rightMargin=50,
        leftMargin=50,
        topMargin=50,
        bottomMargin=50
    )

    elements = []

    if LOGO_PATH and os.path.exists(LOGO_PATH):
        img = Image(LOGO_PATH, width=2.5 * inch, height=1.25 * inch)
        elements.append(img)
        elements.append(Spacer(1, 15))

    elements.extend(flowables)
    doc.build(elements)

# -----------------------------
# EXCEL TRACKER
# -----------------------------
def update_excel(date_folder, appointment_id, entity_type, entity_name, client_key, amount):
    file_path = os.path.join(date_folder, "appointments_log.xlsx")

    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.append(["Appointment Number","Date","Type","Entity Name","Client Key","Amount"])
    else:
        wb = load_workbook(file_path)
        ws = wb.active

    ws.append([
        appointment_id,
        datetime.now().strftime("%Y-%m-%d"),
        entity_type,
        entity_name,
        client_key,
        amount
    ])

    wb.save(file_path)

# -----------------------------
# CREATE FOLDERS (USER-BASED)
# -----------------------------
def setup_folders(client_key):
    date_str = datetime.now().strftime("%Y-%m-%d")

    date_folder = os.path.join(ROOT_APPOINTMENTS_FOLDER, date_str)
    client_folder = os.path.join(date_folder, client_key)

    cloud_folder = os.path.join(client_folder, "Cloud Services")
    consultant_folder = os.path.join(client_folder, "Consultants")

    os.makedirs(cloud_folder, exist_ok=True)
    os.makedirs(consultant_folder, exist_ok=True)

    return client_folder, cloud_folder, consultant_folder

# -----------------------------
# FILE PICKER
# -----------------------------
def select_invoice_file():
    root = tk.Tk()
    root.withdraw()
    return filedialog.askopenfilename(
        title="Select Invoice TXT File",
        filetypes=[("Text files","*.txt")]
    )

# -----------------------------
# USER INPUT + INIT PATHS
# -----------------------------
def initialize_user(username):
    global USER_FOLDER, ROOT_APPOINTMENTS_FOLDER, LOGO_PATH

    USER_FOLDER = os.path.join("users", username)
    ROOT_APPOINTMENTS_FOLDER = os.path.join(USER_FOLDER, "appointments")
    LOGO_PATH = os.path.join(USER_FOLDER, "logo", "logo.png")

# -----------------------------
# MAIN
# -----------------------------
def main(invoice_path):
    client_key = extract_client_key(invoice_path)
    print("Client:", client_key)

    cloud_data, consultant_data = parse_invoice(invoice_path)

    client_folder, cloud_folder, consultant_folder = setup_folders(client_key)

    # ✅ SINGLE RUN TIMESTAMP (NEW APPOINTMENT BATCH)
    run_timestamp = datetime.now()

    cloud_counter = 1
    consultant_counter = 1

    for name, data in cloud_data.items():
        appointment_id = generate_id("cloud", cloud_counter, client_key)

        flowables = build_letter_flowables(
            name,
            data["tasks"],
            data["total"],
            appointment_id,
            "cloud",
            client_key,
            run_timestamp
        )

        save_pdf(cloud_folder, name, appointment_id, client_key, "cloud", flowables)
        update_excel(client_folder, appointment_id, "Cloud Service", name, client_key, data["total"])
        cloud_counter += 1

    for name, data in consultant_data.items():
        appointment_id = generate_id("consultant", consultant_counter, client_key)

        flowables = build_letter_flowables(
            name,
            data["tasks"],
            data["total"],
            appointment_id,
            "consultant",
            client_key,
            run_timestamp
        )

        save_pdf(consultant_folder, name, appointment_id, client_key, "consultant", flowables)
        update_excel(client_folder, appointment_id, "Consultant", name, client_key, data["total"])
        consultant_counter += 1

    print("✅ PDF appointments created successfully.")

# -----------------------------
# AUTO RUN
# -----------------------------
if __name__ == "__main__":
    username = input("Enter username:\n> ").strip()

    if not username:
        print("❌ Username is required.")
        sys.exit()

    initialize_user(username)

    if len(sys.argv) > 1:
        invoice_file = sys.argv[1]
    else:
        print("📂 Select invoice file...")
        invoice_file = select_invoice_file()

    if invoice_file and os.path.exists(invoice_file):
        main(invoice_file)
    else:
        print("❌ No file selected.")