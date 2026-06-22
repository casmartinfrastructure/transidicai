import os
import re
import sys
from datetime import datetime, timedelta
from collections import defaultdict
import tkinter as tk
from tkinter import filedialog

from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Image,
    Table, TableStyle
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_JUSTIFY, TA_CENTER
from reportlab.lib import colors

from openpyxl import Workbook, load_workbook

# -----------------------------
# CONFIG
# -----------------------------
WHATSAPP_NUMBER = "+263 785 261 617"

# -----------------------------
# GLOBAL
# -----------------------------
USER_FOLDER = None
ROOT_APPOINTMENTS_FOLDER = None
LOGO_PATH = None

# -----------------------------
# GENERATE APPOINTMENT ID
# -----------------------------
def generate_id(entity_type, counter, client_key):
    date_str = datetime.now().strftime("%Y%m%d")
    prefix = "CS" if entity_type == "cloud" else "CON"
    return f"{client_key}-APT-{date_str}-{prefix}-{counter:03d}"

# -----------------------------
# EXTRACT CLIENT KEY
# -----------------------------
def extract_client_key(file_path):
    with open(file_path, "r", encoding="utf-8") as f:
        content = f.read()

    match = re.search(r"\b[A-Z]{2}-[A-Z0-9]{6}\b", content)
    if match:
        return match.group(0)

    raise Exception("Encrypted client key (e.g. BW-74B35A) not found in invoice.")

# -----------------------------
# PARSE INVOICE TABLE
# -----------------------------
def parse_invoice(file_path):
    cloud_data = defaultdict(lambda: {"tasks": [], "total": 0.0})
    consultant_data = defaultdict(lambda: {"tasks": [], "total": 0.0})

    with open(file_path, "r", encoding="utf-8") as f:
        lines = f.readlines()

    table_started = False
    current_row = None

    for line in lines:
        if "| Item |" in line:
            table_started = True
            continue

        if not table_started or "|" not in line or line.startswith("+"):
            continue

        cols = [c.strip() for c in line.split("|")[1:-1]]
        if len(cols) < 11:
            continue

        item, description, cloud_service, consultant = cols[0], cols[1], cols[2], cols[4]
        cloud_amt, consultant_amt = cols[7], cols[8]

        if item:
            if current_row:
                try:
                    cloud_value = float(current_row["cloud_amt"].replace("$", ""))
                    consultant_value = float(current_row["consultant_amt"].replace("$", ""))
                except:
                    cloud_value, consultant_value = 0.0, 0.0

                cloud_name = current_row["cloud_service"].strip()
                consultant_name = current_row["consultant"].strip()
                task = " ".join(current_row["description"].split())

                cloud_data[cloud_name]["tasks"].append(task)
                cloud_data[cloud_name]["total"] += cloud_value

                consultant_data[consultant_name]["tasks"].append(task)
                consultant_data[consultant_name]["total"] += consultant_value

            current_row = {
                "description": description,
                "cloud_service": cloud_service,
                "consultant": consultant,
                "cloud_amt": cloud_amt,
                "consultant_amt": consultant_amt,
            }

        elif current_row:
            if description:
                current_row["description"] += " " + description
            if cloud_service:
                current_row["cloud_service"] += " " + cloud_service
            if consultant:
                current_row["consultant"] += " " + consultant

    if current_row:
        try:
            cloud_value = float(current_row["cloud_amt"].replace("$", ""))
            consultant_value = float(current_row["consultant_amt"].replace("$", ""))
        except:
            cloud_value, consultant_value = 0.0, 0.0

        cloud_name = current_row["cloud_service"].strip()
        consultant_name = current_row["consultant"].strip()
        task = " ".join(current_row["description"].split())

        cloud_data[cloud_name]["tasks"].append(task)
        cloud_data[cloud_name]["total"] += cloud_value

        consultant_data[consultant_name]["tasks"].append(task)
        consultant_data[consultant_name]["total"] += consultant_value

    return cloud_data, consultant_data

# -----------------------------
# PARSE HOURS INPUT FILE
# Extract exact entity name -> estimated hours
# -----------------------------
def parse_hours_file(file_path):
    """
    Parses a TXT file with entries like:

    1. Valentine Mujana (Valentine Mujana.pdf)
       Match Score: 0.330
       Experience Level: senior
       Task Complexity: low
       Estimated Hours: 5.68 hrs

    Returns:
        {
            "Valentine Mujana": 5.68,
            "Tinashe Chimusasa": 7.76,
            ...
        }

    Also stores the filename inside brackets without extension
    so cloud services can also match if their invoice name matches
    the filename rather than the visible display name.
    """
    with open(file_path, "r", encoding="utf-8") as f:
        content = f.read()

    lines = content.splitlines()
    hours_map = {}
    current_name = None
    current_file_name = None

    for line in lines:
        stripped = line.strip()

        # Match lines like:
        # 1. Valentine Mujana (Valentine Mujana.pdf)
        name_match = re.match(r"^\d+\.\s+(.*?)\s+\((.*?)\)\s*$", stripped)
        if name_match:
            current_name = name_match.group(1).strip()
            current_file_name = os.path.splitext(name_match.group(2).strip())[0].strip()
            continue

        # Match lines like:
        # Estimated Hours: 5.68 hrs
        hours_match = re.match(r"^Estimated Hours:\s*([\d.]+)\s*hrs?$", stripped, re.IGNORECASE)
        if hours_match and current_name:
            try:
                hours_value = float(hours_match.group(1))
            except:
                hours_value = 0.0

            hours_map[current_name] = hours_value

            if current_file_name:
                hours_map[current_file_name] = hours_value

            current_name = None
            current_file_name = None

    return hours_map

# -----------------------------
# APPLY HOURS MULTIPLIER
# Multiplies invoice totals by exact matched hours
# -----------------------------
def apply_hours_multiplier(entity_data, hours_map):
    """
    entity_data format:
        {
            "Name A": {"tasks": [...], "total": 100.0},
            "Name B": {"tasks": [...], "total": 200.0}
        }

    If exact name exists in hours_map:
        new_total = old_total * hours

    If no match found:
        total remains unchanged
    """
    for entity_name, data in entity_data.items():
        if entity_name in hours_map:
            data["total"] = data["total"] * hours_map[entity_name]
    return entity_data

# -----------------------------
# FORMAL LETTER BUILDER (HEADER REMOVED)
# -----------------------------
def build_letter_flowables(entity_name, tasks, total_cost, appointment_id,
                           entity_type, client_key, appointment_time):

    styles = getSampleStyleSheet()

    heading_style = ParagraphStyle(
        "heading",
        parent=styles["Heading3"],
        spaceBefore=10,
        spaceAfter=6,
        textColor=colors.black
    )

    body_style = ParagraphStyle(
        "body",
        parent=styles["Normal"],
        spaceAfter=8,
        leading=16,
        alignment=TA_JUSTIFY,
        fontSize=10.5
    )

    footer_style = ParagraphStyle(
        "footer",
        parent=styles["Normal"],
        alignment=TA_CENTER,
        fontSize=9,
        textColor=colors.grey
    )

    entity_label = "Cloud Service" if entity_type == "cloud" else "Consultant"
    elements = []

    deadline = appointment_time + timedelta(minutes=15)

    # -----------------------------
    # REFERENCE TABLE (FORMAL BLOCK)
    # -----------------------------
    ref_table = [
        ["Client Key", client_key],
        ["Appointment Reference", appointment_id],
        ["Appointed Entity", entity_name],
        ["Entity Type", entity_label],
        ["Appointment Time", appointment_time.strftime("%Y-%m-%d %H:%M:%S")],
        ["Response Deadline", deadline.strftime("%Y-%m-%d %H:%M:%S")]
    ]

    table = Table(ref_table, colWidths=[180, 300])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.lightgrey),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("PADDING", (0, 0), (-1, -1), 6),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 12))

    # -----------------------------
    # BODY
    # -----------------------------
    elements.append(Paragraph(f"Dear {entity_name},", heading_style))
    elements.append(Paragraph("Appointment Confirmation", heading_style))

    appointment_text = f"""
    In accordance with the On-Demand Procurement Platform Participation Agreement, we are pleased to notify you that you have been selected for the Procurement Request(s) listed below.<br/><br/>
    Kindly accept or decline this appointment within <b>15 minutes</b> of receiving this notification.<br/><br/>
    <b>Acceptance:</b> Further instructions will be issued.<br/>
    <b>Decline:</b> No action required.<br/>
    <b>No Response:</b> Automatically treated as declined.<br/><br/>
    <b>Payment Terms:</b> The stated amount represents the total payable for task completion and excludes operational costs unless otherwise agreed in writing.<br/><br/>
    All services are executed under the Transidic Startup Studio Service Agreement.
    """

    elements.append(Paragraph(appointment_text, body_style))

    elements.append(Spacer(1, 10))
    elements.append(Paragraph("Procurement Request(s)", heading_style))

    for idx, task in enumerate(tasks, start=1):
        elements.append(Paragraph(f"{idx}. {task}", body_style))

    elements.append(Spacer(1, 10))
    elements.append(Paragraph(f"Total Offer Amount: <b>${total_cost:.2f}</b>", heading_style))

    elements.append(Spacer(1, 15))

    # -----------------------------
    # SIGNATURE
    # -----------------------------
    elements.append(Paragraph("Yours faithfully,", body_style))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph("Transidic Startup Studio Private Limited", body_style))
    elements.append(Paragraph("WhatsApp: +263 785 261 617", footer_style))
    elements.append(Paragraph("Email: transidicstudio@gmail.com", footer_style))

    return elements

# -----------------------------
# SAVE PDF (TOP LOGO REMOVED)
# -----------------------------
def save_pdf(folder, name, appointment_id, client_key, entity_type, flowables):
    safe_name = re.sub(r"[^\w\- ]", "", name).replace(" ", "_")
    safe_type = "Cloud_Service" if entity_type == "cloud" else "Consultant"

    file_path = os.path.join(folder, f"{appointment_id}_{safe_type}_{safe_name}.pdf")

    doc = SimpleDocTemplate(
        file_path,
        pagesize=A4,
        rightMargin=50,
        leftMargin=50,
        topMargin=50,
        bottomMargin=50
    )

    elements = []
    elements.extend(flowables)

    doc.build(elements)

# -----------------------------
# EXCEL TRACKER
# -----------------------------
def update_excel(date_folder, appointment_id, entity_type, entity_name, client_key, amount):
    file_path = os.path.join(date_folder, "appointments_log.xlsx")

    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.append(["Appointment Number","Date","Type","Entity Name","Client Key","Amount"])
    else:
        wb = load_workbook(file_path)
        ws = wb.active

    ws.append([
        appointment_id,
        datetime.now().strftime("%Y-%m-%d"),
        entity_type,
        entity_name,
        client_key,
        amount
    ])

    wb.save(file_path)

# -----------------------------
# FOLDERS
# -----------------------------
def setup_folders(client_key):
    date_str = datetime.now().strftime("%Y-%m-%d")

    date_folder = os.path.join(ROOT_APPOINTMENTS_FOLDER, date_str)
    client_folder = os.path.join(date_folder, client_key)

    cloud_folder = os.path.join(client_folder, "Cloud Services")
    consultant_folder = os.path.join(client_folder, "Consultants")

    os.makedirs(cloud_folder, exist_ok=True)
    os.makedirs(consultant_folder, exist_ok=True)

    return client_folder, cloud_folder, consultant_folder

# -----------------------------
# FILE PICKERS
# -----------------------------
def select_invoice_file():
    root = tk.Tk()
    root.withdraw()
    return filedialog.askopenfilename(
        title="Select Invoice TXT File",
        filetypes=[("Text files","*.txt")]
    )

def select_hours_file():
    root = tk.Tk()
    root.withdraw()
    return filedialog.askopenfilename(
        title="Select Hours/Matching TXT File",
        filetypes=[("Text files","*.txt")]
    )

# -----------------------------
# USER INIT
# -----------------------------
def initialize_user(username):
    global USER_FOLDER, ROOT_APPOINTMENTS_FOLDER, LOGO_PATH

    USER_FOLDER = os.path.join("users", username)
    ROOT_APPOINTMENTS_FOLDER = os.path.join(USER_FOLDER, "appointments")
    LOGO_PATH = os.path.join(USER_FOLDER, "logo", "logo.png")

# -----------------------------
# MAIN
# -----------------------------
def main(invoice_path, hours_file_path=None):
    client_key = extract_client_key(invoice_path)
    print("Client:", client_key)

    cloud_data, consultant_data = parse_invoice(invoice_path)

    # -----------------------------------------
    # NEW: Parse hours file and multiply totals
    # -----------------------------------------
    if hours_file_path and os.path.exists(hours_file_path):
        hours_map = parse_hours_file(hours_file_path)

        # Apply exact-name matching to BOTH cloud and consultant offers
        cloud_data = apply_hours_multiplier(cloud_data, hours_map)
        consultant_data = apply_hours_multiplier(consultant_data, hours_map)

    client_folder, cloud_folder, consultant_folder = setup_folders(client_key)

    run_timestamp = datetime.now()

    cloud_counter = 1
    consultant_counter = 1

    for name, data in cloud_data.items():
        appointment_id = generate_id("cloud", cloud_counter, client_key)

        flowables = build_letter_flowables(
            name,
            data["tasks"],
            data["total"],
            appointment_id,
            "cloud",
            client_key,
            run_timestamp
        )

        save_pdf(cloud_folder, name, appointment_id, client_key, "cloud", flowables)
        update_excel(client_folder, appointment_id, "Cloud Service", name, client_key, data["total"])
        cloud_counter += 1

    for name, data in consultant_data.items():
        appointment_id = generate_id("consultant", consultant_counter, client_key)

        flowables = build_letter_flowables(
            name,
            data["tasks"],
            data["total"],
            appointment_id,
            "consultant",
            client_key,
            run_timestamp
        )

        save_pdf(consultant_folder, name, appointment_id, client_key, "consultant", flowables)
        update_excel(client_folder, appointment_id, "Consultant", name, client_key, data["total"])
        consultant_counter += 1

    print("✅ PDF appointments created successfully.")

# -----------------------------
# AUTO RUN
# -----------------------------
if __name__ == "__main__":
    username = input("Enter username:\n> ").strip()

    if not username:
        print("❌ Username is required.")
        sys.exit()

    initialize_user(username)

    # Optional CLI usage:
    # python script.py invoice.txt hours.txt
    if len(sys.argv) > 1:
        invoice_file = sys.argv[1]
    else:
        print("📂 Select invoice file...")
        invoice_file = select_invoice_file()

    if not invoice_file or not os.path.exists(invoice_file):
        print("❌ No invoice file selected.")
        sys.exit()

    if len(sys.argv) > 2:
        hours_file = sys.argv[2]
    else:
        print("📂 Select hours/matching file...")
        hours_file = select_hours_file()

    if hours_file and os.path.exists(hours_file):
        main(invoice_file, hours_file)
    else:
        # Still runs if hours file not selected, preserving existing behavior
        print("⚠️ No hours file selected. Running with original invoice totals.")
        main(invoice_file, None)